import os
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries
import json
from datetime import datetime
from .skip_pipeline import should_translate

def extract_excel_content_to_json(file_path):
    workbook = load_workbook(file_path)
    cell_data = []
    count = 0

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        merged_cells_ranges = sheet.merged_cells.ranges

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None or isinstance(cell.value, datetime) or not should_translate(str(cell.value)):
                    continue
                if isinstance(cell, MergedCell):
                    continue
                is_merged_cell = False
                for merged_range in merged_cells_ranges:
                    min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                    if cell.row == min_row and cell.column == min_col:
                        is_merged_cell = True
                        break
                # Convert datetime values to string
                cell_value = str(cell.value).replace("\n", "␊").replace("\r", "␍")
                if isinstance(cell_value, datetime):
                    cell_value = cell_value.isoformat()
                count += 1
                cell_info = {
                    "count": count,
                    "sheet": sheet_name,
                    "row": cell.row,
                    "column": cell.column,
                    "value": cell_value,  # Corrected to use the processed value
                    "is_merged": is_merged_cell,
                }
                cell_data.append(cell_info)
    temp_folder = os.path.join('temp')
    if not os.path.exists(temp_folder):
        os.makedirs(temp_folder)
    json_path = os.path.join(temp_folder, "src.json")
    with open(json_path, "w", encoding="utf-8") as json_file:
        json.dump(cell_data, json_file, ensure_ascii=False, indent=4)

    return json_path

def modify_json(data_list):
    combined_data = {}
    for entry in data_list:
        if entry.startswith("```json"):
            entry = entry[len("```json"):].strip()
        if entry.endswith("```"):
            entry = entry[:-len("```")].strip()
        
        try:
            json_data = json.loads(entry)
            if isinstance(json_data, dict):
                combined_data.update(json_data)
        except json.JSONDecodeError:
            print(f"Warning: Skipping invalid JSON entry: {entry}")
            continue
    
    return combined_data

def write_translated_content_to_excel(file_path, original_json_path, translated_json_path):
    workbook = load_workbook(file_path)
    with open(original_json_path, "r", encoding="utf-8") as original_file:
        original_data = json.load(original_file)
    with open(translated_json_path, "r", encoding="utf-8") as translated_file:
        translated_raw = json.load(translated_file)
        translated_data = modify_json(translated_raw)

    # Write translated content back to Excel
    for cell_info in original_data:
        count = cell_info["count"]
        sheet_name = cell_info["sheet"]
        row = cell_info["row"]
        column = cell_info["column"]
        is_merged = cell_info.get("is_merged", False)

        # Get translated value
        value = translated_data.get(str(count), "")
        value = value.replace("␊", "\n").replace("␍", "\r")
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found in workbook.")
            continue

        # Write translated value to cell
        sheet = workbook[sheet_name]
        cell = sheet.cell(row=row, column=column)
        cell.value = value

        # Handle merged cells if necessary
        if is_merged:
            sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column)

    # Save file with "_translated" suffix
    result_folder = os.path.join('result')
    if not os.path.exists(result_folder):
        os.makedirs(result_folder)
    result_path = os.path.join(result_folder, f"{os.path.splitext(os.path.basename(file_path))[0]}_translated{os.path.splitext(file_path)[1]}")
    workbook.save(result_path)
    return result_path
    
if __name__=="__main__":
    file_path = "test.xlsx"
    with open(file_path, "rb") as file:
        result_json_path = extract_cells_to_json(file)