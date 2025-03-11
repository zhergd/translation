import os
from lxml import etree
import zipfile
import json
from .skip_pipeline import should_translate
from config.log_config import app_logger


def extract_excel_content_to_json(file_path):
    """
    Extract content from Excel file using lxml for better performance.
    """
    # Excel files are zip files containing XML
    cell_data = []
    count = 0
    
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        # First, load shared strings if available
        shared_strings = []
        if 'xl/sharedStrings.xml' in zip_ref.namelist():
            with zip_ref.open('xl/sharedStrings.xml') as shared_strings_file:
                xml_content = shared_strings_file.read()
                root = etree.fromstring(xml_content)
                # XML namespace
                ns = {'xmlns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                
                # Extract all shared strings - try different possible structures
                for si in root.xpath("//xmlns:si", namespaces=ns):
                    # Try to get text from t elements
                    text_parts = []
                    for t in si.xpath(".//xmlns:t", namespaces=ns):
                        if t.text is not None:  # Check if text is not None
                            text_parts.append(t.text)
                    shared_strings.append(''.join(text_parts))
                
                app_logger.info(f"Loaded {len(shared_strings)} shared strings")
        
        # Get relationship info to map sheet IDs to file paths
        rels = {}
        if 'xl/_rels/workbook.xml.rels' in zip_ref.namelist():
            try:
                with zip_ref.open('xl/_rels/workbook.xml.rels') as rels_file:
                    rels_content = rels_file.read()
                    rels_root = etree.fromstring(rels_content)
                    
                    for rel in rels_root.xpath("//Relationship"):
                        rel_id = rel.get('Id')
                        target = rel.get('Target')
                        rels[rel_id] = target
                    
                    app_logger.info(f"Loaded relationships: {rels}")
            except Exception as e:
                app_logger.error(f"Error reading relationships: {str(e)}")
        
        # Get workbook.xml to read sheet names
        sheet_to_file_map = {}
        try:
            with zip_ref.open('xl/workbook.xml') as workbook_file:
                workbook_content = workbook_file.read()
                workbook_root = etree.fromstring(workbook_content)
                ns = {'xmlns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                     'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
                
                # Map sheet names to file paths
                for i, sheet in enumerate(workbook_root.xpath("//xmlns:sheet", namespaces=ns)):
                    sheet_name = sheet.get('name')
                    r_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                    
                    # Find the file path using the relationship ID
                    if r_id in rels:
                        target = rels[r_id]
                        sheet_file = 'xl/' + target.lstrip('/')  # Ensure proper path
                    else:
                        # Fallback: use sheet index to construct file path
                        sheet_file = f'xl/worksheets/sheet{i+1}.xml'
                    
                    sheet_to_file_map[sheet_name] = sheet_file
                
                app_logger.info(f"Sheet to file mapping: {sheet_to_file_map}")
        except Exception as e:
            app_logger.error(f"Error reading workbook.xml: {str(e)}")
            # Fallback: Match files by index
            sheet_files = sorted([f for f in zip_ref.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')])
            sheet_to_file_map = {f"Sheet{i+1}": file for i, file in enumerate(sheet_files)}
            app_logger.info(f"Using fallback sheet mapping: {sheet_to_file_map}")
        
        # Process each sheet using the mapping
        for sheet_name, sheet_file in sheet_to_file_map.items():
            if sheet_file not in zip_ref.namelist():
                app_logger.warning(f"Sheet file {sheet_file} for '{sheet_name}' not found")
                continue
            
            app_logger.info(f"Processing sheet: {sheet_name}, file: {sheet_file}")
            
            try:
                with zip_ref.open(sheet_file) as sheet_xml:
                    xml_content = sheet_xml.read()
                    root = etree.fromstring(xml_content)
                    
                    # XML namespace
                    ns = {'xmlns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    
                    # Get merged cells information
                    merged_cells = []
                    merge_cells_elem = root.xpath("//xmlns:mergeCells/xmlns:mergeCell", namespaces=ns)
                    for merged_cell_range in merge_cells_elem:
                        merged_cells.append(merged_cell_range.get('ref'))
                    
                    app_logger.info(f"Found {len(merged_cells)} merged cell ranges")
                    
                    # Function to check if a cell is the top-left of a merged range
                    def is_merge_start(cell_ref):
                        for merge_range in merged_cells:
                            parts = merge_range.split(':')
                            if len(parts) == 2 and parts[0] == cell_ref:
                                return True
                        return False
                    
                    # Process all rows and cells
                    row_elements = root.xpath("//xmlns:row", namespaces=ns)
                    app_logger.info(f"Found {len(row_elements)} rows in {sheet_name}")
                    
                    for row in row_elements:
                        row_index = int(row.get('r', '0'))
                        
                        cell_elements = row.xpath(".//xmlns:c", namespaces=ns)
                        
                        for cell in cell_elements:
                            cell_ref = cell.get('r')  # Cell reference like A1, B5, etc.
                            cell_type = cell.get('t')  # Cell type (s=shared string, inlineStr, b=boolean, etc.)
                            
                            # Get cell value
                            value = None
                            
                            if cell_type == 's':
                                # Shared string
                                value_element = cell.find(".//xmlns:v", namespaces=ns)
                                if value_element is not None and value_element.text:
                                    try:
                                        string_index = int(value_element.text)
                                        if 0 <= string_index < len(shared_strings):
                                            value = shared_strings[string_index]
                                    except (ValueError, IndexError) as e:
                                        app_logger.error(f"Error with shared string index in {cell_ref}: {str(e)}")
                            
                            elif cell_type == 'inlineStr':
                                # Inline string
                                is_element = cell.find(".//xmlns:is", namespaces=ns)
                                if is_element is not None:
                                    t_elements = is_element.xpath(".//xmlns:t", namespaces=ns)
                                    value = ''.join(t.text for t in t_elements if t.text)
                            
                            elif cell_type == 'str':
                                # String formula result
                                value_element = cell.find(".//xmlns:v", namespaces=ns)
                                if value_element is not None:
                                    value = value_element.text
                            
                            else:
                                # Other types (number, boolean, etc.)
                                value_element = cell.find(".//xmlns:v", namespaces=ns)
                                if value_element is not None and value_element.text:
                                    value = value_element.text
                            
                            # Skip empty cells or non-translatable content
                            if value is None:
                                continue
                                
                            # Check if we should translate this content
                            if not should_translate(str(value)):
                                continue
                            
                            # Clean and prepare value
                            cell_value = str(value).replace("\n", "␊").replace("\r", "␍")
                            
                            # Extract column from cell reference (e.g., A1 -> column=1)
                            import re
                            col_match = re.match(r'([A-Za-z]+)\d+', cell_ref)
                            if col_match:
                                col_str = col_match.group(1)
                                col_num = 0
                                for c in col_str:
                                    col_num = col_num * 26 + (ord(c.upper()) - ord('A') + 1)
                            else:
                                app_logger.warning(f"Could not parse column from {cell_ref}")
                                continue
                            
                            # Check if this is a merged cell start
                            is_merged = is_merge_start(cell_ref)
                            
                            count += 1
                            cell_info = {
                                "count": count,
                                "sheet": sheet_name,
                                "row": row_index,
                                "column": col_num,
                                "cell_ref": cell_ref,
                                "value": cell_value,
                                "is_merged": is_merged,
                            }
                            cell_data.append(cell_info)
                                   
            except Exception as e:
                app_logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                import traceback
                app_logger.error(traceback.format_exc())
    
    app_logger.info(f"Total cells extracted: {count}")
    
    # Save the extracted data to JSON
    filename = os.path.splitext(os.path.basename(file_path))[0]
    temp_folder = os.path.join("temp", filename)
    os.makedirs(temp_folder, exist_ok=True)
    json_path = os.path.join(temp_folder, "src.json")
    
    with open(json_path, "w", encoding="utf-8") as json_file:
        json.dump(cell_data, json_file, ensure_ascii=False, indent=4)

    return json_path


def write_translated_content_to_excel(file_path, original_json_path, translated_json_path):
    """
    Write translated content back to the Excel file.
    Falling back to openpyxl as it's more reliable for writing Excel files.
    """
    from openpyxl import load_workbook
    
    workbook = load_workbook(file_path)

    # Load original and translated JSON data
    with open(original_json_path, "r", encoding="utf-8") as original_file:
        original_data = json.load(original_file)
    
    with open(translated_json_path, "r", encoding="utf-8") as translated_file:
        translated_data = json.load(translated_file)

    # Convert translations to a dictionary {count: translated_value}
    translations = {str(item["count"]): item["translated"] for item in translated_data}

    # Write translated content back to Excel
    cells_updated = 0
    for cell_info in original_data:
        count = str(cell_info["count"])
        sheet_name = cell_info["sheet"]
        row = cell_info["row"]
        column = cell_info["column"]
        original_text = cell_info["value"]
        
        # Skip if sheet doesn't exist
        if sheet_name not in workbook.sheetnames:
            app_logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            continue
            
        # Get translated text
        value = translations.get(count)
        if value is None:
            app_logger.warning(
                f"Translation missing for count {count}. Original text: '{original_text}'"
            )
            continue
        
        # Replace special characters back
        value = value.replace("␊", "\n").replace("␍", "\r")

        # Write to the Excel cell
        sheet = workbook[sheet_name]
        try:
            cell = sheet.cell(row=row, column=column)
            cell.value = value
            cells_updated += 1
        except Exception as e:
            app_logger.error(f"Error writing to cell row={row}, column={column}: {str(e)}")

    app_logger.info(f"Updated {cells_updated} cells with translations")
    
    # Save the modified Excel file
    result_folder = os.path.join('result')
    os.makedirs(result_folder, exist_ok=True)
    
    result_path = os.path.join(
        result_folder,
        f"{os.path.splitext(os.path.basename(file_path))[0]}_translated{os.path.splitext(file_path)[1]}"
    )
    
    workbook.save(result_path)
    app_logger.info(f"Translated Excel saved to: {result_path}")
    return result_path